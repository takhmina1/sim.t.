# import datetime
# from openpyxl import Workbook
# from robots.models import Robot

# def generate_excel_summary():
#     """
#     Генерация Excel-файла со сводкой по произведённым роботам за последнюю неделю.
#     """
#     # Определяем временной диапазон - последние 7 дней
#     today = datetime.date.today()
#     week_ago = today - datetime.timedelta(days=7)

#     # Получаем роботов, произведённых за последние 7 дней
#     robots = Robot.objects.filter(created_at__date__range=[week_ago, today])

#     # Создаём Excel-файл
#     workbook = Workbook()

#     # Группировка по моделям
#     models = robots.values_list('model', flat=True).distinct()
#     for model in models:
#         sheet = workbook.create_sheet(title=model)
#         sheet.append(["Модель", "Версия", "Количество за неделю"])

#         # Группировка по версиям
#         versions = robots.filter(model=model).values('version').distinct()
#         for version in versions:
#             count = robots.filter(model=model, version=version['version']).count()
#             sheet.append([model, version['version'], count])

#     # Удаляем стандартный пустой лист
#     if "Sheet" in workbook.sheetnames:
#         workbook.remove(workbook["Sheet"])

#     return workbook




import datetime
from openpyxl import Workbook
from robots.models import Robot

def generate_excel_summary():
    """
    Генерация Excel-файла со сводкой по произведённым роб
            send_mail(
                'Ваш заказ на робота в наличии!',
                f'Добрый день!\n\nНедавно отам за последнюю неделю.
    """
    # Определяем временной диапазон - последние 7 дней
    today = datetime.date.today()
    week_ago = today - datetime.timedelta(days=7)

    # Получаем роботов, произведённых за последние 7 дней
    robots = Robot.objects.filter(created_at__date__range=[week_ago, today])

    if not robots.exists():
        # Если роботов нет, возвращаем пустой Excel файл с одним листом
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Нет данных", "Нет данных", "Нет данных"])
        return workbook

    # Создаём Excel-файл
    workbook = Workbook()

    # Группировка по моделям
    models = robots.values_list('model', flat=True).distinct()
    for model in models:
        sheet = workbook.create_sheet(title=model)
        sheet.append(["Модель", "Версия", "Количество за неделю"])

        # Группировка по версиям
        versions = robots.filter(model=model).values('version').distinct()
        for version in versions:
            count = robots.filter(model=model, version=version['version']).count()
            sheet.append([model, version['version'], count])

    # Удаляем стандартный пустой лист, если он есть
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    return workbook
